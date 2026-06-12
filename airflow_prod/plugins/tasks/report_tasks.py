# airflow_prod/plugins/tasks/report_tasks.py
# Funkcje do przygotowania raportu pipeline'u
# Generuje pliki Excel i txt do wysłania emailem
#
# Pliki generowane w data/reports/:
#   report_warnings_{date}.xlsx       ← zakładki z widokami WARNING (tylko gdy są dane)
#   report_details_{date}.xlsx        ← kroki pipeline + przestrzeń dyskowa (zawsze)
#   error_scrape_{portal}_{date}.txt  ← błędy scraperów (tylko gdy były błędy)
#   error_details_{date}.txt          ← błąd krytyczny (tylko przy ERROR)

import os
import glob
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import get_db_connection
from config import PROJECT_DIR

REPORTS_DIR = os.path.join(PROJECT_DIR, "data", "reports")


# ── Funkcje pomocnicze ───────────────────────────────────────────────────────

def _format_duration(seconds) -> str:
    """Formatuje sekundy do formatu hh:mm:ss."""
    if not seconds:
        return "N/A"
    try:
        seconds = int(seconds)
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f"{h:02d}:{m:02d}:{s:02d}"
    except Exception:
        return "N/A"


def _strip_timezone(value):
    """Usuwa informację o strefie czasowej z obiektu datetime dla Excel."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _safe_fetch(conn, sql: str, params: tuple = None) -> tuple:
    """
    Bezpieczne wykonanie zapytania SQL.
    Zwraca (columns, rows).
    W przypadku błędu lub braku danych zwraca ([], []).
    Robi rollback po błędzie żeby nie blokować transakcji.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            columns = [desc[0] for desc in cur.description]
            rows    = cur.fetchall()
            return columns, rows
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"[report] BŁĄD zapytania SQL: {e}")
        return [], []


def _safe_fetch_one(conn, sql: str, params: tuple = None):
    """
    Bezpieczne wykonanie zapytania SQL zwracającego jeden wiersz.
    Zwraca wiersz lub None w przypadku błędu/braku danych.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            return cur.fetchone()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"[report] BŁĄD zapytania SQL: {e}")
        return None


def _safe_fetch_scalar(conn, sql: str, params: tuple = None, default=0):
    """
    Bezpieczne wykonanie zapytania SQL zwracającego jedną wartość skalarną.
    Zwraca wartość lub default w przypadku błędu/braku danych.
    """
    row = _safe_fetch_one(conn, sql, params)
    if row is None or len(row) == 0:
        return default
    return row[0] if row[0] is not None else default


# ── Style Excel ──────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ERROR_FILL   = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SKIPPED_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

STATUS_FILLS = {
    "SUCCESS": SUCCESS_FILL,
    "ERROR":   ERROR_FILL,
    "SKIPPED": SKIPPED_FILL,
    "PENDING": WARNING_FILL,
}

# ── Widoki WARNING — zakładki w report_warnings.xlsx ────────────────────────
WARNING_VIEWS = [
    # (schema, view_name, tab_title)
    ("bronze", "v_test_data_raw_offers",     "bronze_raw_offers_check"),
    ("bronze", "v_test_double_raw_offers",   "bronze_double_offers"),
    ("silver", "v_test_data_offer_benefits", "silver_benefits"),
    ("silver", "v_test_data_offer_contracts","silver_contracts"),
    ("silver", "v_test_data_offer_levels",   "silver_levels"),
    ("silver", "v_test_data_offer_modes",    "silver_modes"),
    ("silver", "v_test_data_offer_salaries", "silver_salaries"),
    ("silver", "v_test_data_offer_schedules","silver_schedules"),
    ("silver", "v_test_data_offer_specs",    "silver_specs"),
    ("silver", "v_test_data_offer_tech",     "silver_tech"),
]


def _style_header_row(ws, columns, start_row: int = 1):
    """Styluje wiersz nagłówka w arkuszu Excel."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_column_width(ws):
    """Automatycznie dopasowuje szerokość kolumn."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


# ── Czyszczenie folderu reports ──────────────────────────────────────────────

def _cleanup_reports_dir():
    """Czyści folder data/reports/ przed wygenerowaniem nowego raportu."""
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)
        print(f"[report] Utworzono folder: {REPORTS_DIR}")
        return
    files = glob.glob(os.path.join(REPORTS_DIR, "*"))
    for f in files:
        os.remove(f)
    print(f"[report] Wyczyszczono {len(files)} plików z {REPORTS_DIR}")


# ── Generowanie report_warnings.xlsx ────────────────────────────────────────

def _generate_warnings_excel(run_date: str, conn) -> str | None:
    """
    Generuje report_warnings_{date}.xlsx z zakładkami per widok WARNING.
    Zakładka tworzona tylko gdy widok zwróci dane.
    Jeśli żaden widok nie zwróci danych — plik NIE jest tworzony, zwraca None.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets_created = 0

    for schema, view_name, tab_title in WARNING_VIEWS:
        columns, rows = _safe_fetch(conn, f"SELECT * FROM {schema}.{view_name} LIMIT 10000;")

        if not rows:
            print(f"[report] {schema}.{view_name} — brak danych, pomijam zakładkę.")
            continue

        ws = wb.create_sheet(title=tab_title)
        _style_header_row(ws, columns)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=_strip_timezone(value))

        _auto_column_width(ws)
        sheets_created += 1
        print(f"[report] {schema}.{view_name} → zakładka '{tab_title}' ({len(rows)} wierszy)")

    if sheets_created == 0:
        print("[report] Brak danych w widokach WARNING — nie tworzę report_warnings.xlsx")
        return None

    file_path = os.path.join(REPORTS_DIR, f"report_warnings_{run_date}.xlsx")
    wb.save(file_path)
    print(f"[report] Zapisano: {file_path} ({sheets_created} zakładek)")
    return file_path


# ── Generowanie report_details.xlsx ─────────────────────────────────────────

def _generate_details_excel(run_date: str, pipeline_run_id: int, conn) -> str:
    """
    Generuje report_details_{date}.xlsx z dwiema zakładkami:
    - pipeline_steps: kroki pipeline z czasami i statusami + podsumowanie
    - disk_usage: przestrzeń dyskowa per schema
    Plik generowany zawsze.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Zakładka 1: pipeline_steps ───────────────────────────────────────────
    ws_steps = wb.create_sheet(title="pipeline_steps")

    columns, rows = _safe_fetch(conn,
        """
        SELECT
            s.step_name,
            TO_CHAR(s.start_time AT TIME ZONE 'Europe/Warsaw', 'YYYY-MM-DD HH24:MI:SS') AS start_time,
            TO_CHAR(s.end_time   AT TIME ZONE 'Europe/Warsaw', 'YYYY-MM-DD HH24:MI:SS') AS end_time,
            CASE
                WHEN s.end_time IS NOT NULL
                THEN EXTRACT(EPOCH FROM (s.end_time - s.start_time))::INT
            END AS duration_seconds,
            s.status
        FROM maintenance.pipeline_run_step s
        WHERE s.pipeline_run_id = %s
        ORDER BY s.start_time;
        """,
        (pipeline_run_id,)
    )

    if columns:
        _style_header_row(ws_steps, columns)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws_steps.cell(row=row_idx, column=col_idx, value=_strip_timezone(value))
            # Koloruj wiersz wg statusu (status jest w indeksie 4)
            status = row[4] if len(row) > 4 else None
            if status in STATUS_FILLS:
                for col_idx in range(1, len(columns) + 1):
                    ws_steps.cell(row=row_idx, column=col_idx).fill = STATUS_FILLS[status]
    else:
        ws_steps.cell(row=1, column=1, value="Brak danych o krokach pipeline.")

    # Podsumowanie na dole — tylko gdy jest rekord pipeline_run
    run_summary = _safe_fetch_one(conn,
        """
        SELECT
            TO_CHAR(start_time AT TIME ZONE 'Europe/Warsaw', 'YYYY-MM-DD HH24:MI:SS') AS start_time,
            TO_CHAR(end_time   AT TIME ZONE 'Europe/Warsaw', 'YYYY-MM-DD HH24:MI:SS') AS end_time,
            EXTRACT(EPOCH FROM (end_time - start_time))::INT AS duration_seconds,
            status
        FROM maintenance.pipeline_run
        WHERE id = %s;
        """,
        (pipeline_run_id,)
    )

    if run_summary and len(run_summary) >= 4:
        summary_row = len(rows) + 3
        ws_steps.cell(row=summary_row,     column=1, value="PODSUMOWANIE").font = Font(bold=True)
        ws_steps.cell(row=summary_row + 1, column=1, value="Start:")
        ws_steps.cell(row=summary_row + 1, column=2, value=str(run_summary[0]) if run_summary[0] else "N/A")
        ws_steps.cell(row=summary_row + 2, column=1, value="Koniec:")
        ws_steps.cell(row=summary_row + 2, column=2, value=str(run_summary[1]) if run_summary[1] else "N/A")
        ws_steps.cell(row=summary_row + 3, column=1, value="Czas całkowity:")
        ws_steps.cell(row=summary_row + 3, column=2, value=_format_duration(run_summary[2]))
        ws_steps.cell(row=summary_row + 4, column=1, value="Status:")
        ws_steps.cell(row=summary_row + 4, column=2, value=run_summary[3] if run_summary[3] else "N/A")

    _auto_column_width(ws_steps)

    # ── Zakładka 2: disk_usage ───────────────────────────────────────────────
    ws_disk = wb.create_sheet(title="disk_usage")

    disk_views = [
        ("maintenance", "v_bronze_obj_sizes"),
        ("maintenance", "v_silver_obj_sizes"),
        ("maintenance", "v_gold_obj_sizes"),
        ("maintenance", "v_total_obj_sizes"),
    ]

    current_row  = 1
    any_disk_data = False

    for schema, view_name in disk_views:
        # Wybierz tylko potrzebne kolumny — bez *_num i sort_size
        if view_name == "v_total_obj_sizes":
            sql = f"SELECT schema, total_size, table_size, indexes_size FROM {schema}.{view_name};"
        else:
            sql = f"SELECT table_name, total_size, table_size, indexes_size FROM {schema}.{view_name};"
        columns, rows = _safe_fetch(conn, sql)
        if not rows:
            print(f"[report] {schema}.{view_name} — brak danych dyskowych, pomijam.")
            continue

        any_disk_data = True

        # Nagłówek sekcji
        section_cell = ws_disk.cell(row=current_row, column=1, value=f"{schema}.{view_name}")
        section_cell.font = Font(bold=True)
        current_row += 1

        # Nagłówki kolumn
        _style_header_row(ws_disk, columns, start_row=current_row)
        current_row += 1

        # Dane
        for row in rows:
            for col_idx, value in enumerate(row, 1):
                ws_disk.cell(row=current_row, column=col_idx, value=_strip_timezone(value))
            current_row += 1

        current_row += 1  # odstęp między sekcjami

    if not any_disk_data:
        ws_disk.cell(row=1, column=1, value="Brak danych o przestrzeni dyskowej.")

    _auto_column_width(ws_disk)

    file_path = os.path.join(REPORTS_DIR, f"report_details_{run_date}.xlsx")
    wb.save(file_path)
    print(f"[report] Zapisano: {file_path}")
    return file_path


# ── Generowanie plików txt z błędami scraperów ───────────────────────────────

def _generate_scraper_error_files(run_date: str, pipeline_run_id: int, conn) -> list:
    """
    Generuje osobne pliki txt dla każdego scrapera który się wywalił.
    Jeśli nie było błędów scraperów — zwraca pustą listę.
    """
    columns, failed_scrapers = _safe_fetch(conn,
        """
        SELECT step_name, error_details
        FROM maintenance.pipeline_run_step
        WHERE pipeline_run_id = %s
          AND step_name LIKE 'scrape_%%'
          AND status = 'ERROR'
        ORDER BY step_name;
        """,
        (pipeline_run_id,)
    )

    # Brak błędów scraperów — zwróć pustą listę
    if not failed_scrapers:
        print("[report] Brak błędów scraperów — nie tworzę plików txt.")
        return []

    error_files = []
    for row in failed_scrapers:
        # Zabezpieczenie na wypadek niepełnego wiersza
        if not row or len(row) < 2:
            continue

        step_name     = row[0]
        error_details = row[1]
        portal        = step_name.replace("scrape_", "")
        file_path     = os.path.join(REPORTS_DIR, f"error_{step_name}_{run_date}.txt")

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(f"=== BŁĄD SCRAPERA: {step_name} ===\n")
            f.write(f"Data: {run_date}\n")
            f.write(f"Portal: {portal}\n\n")
            f.write(error_details or "Brak szczegółów błędu.")

        error_files.append(file_path)
        print(f"[report] Zapisano błąd scrapera: {file_path}")

    return error_files


# ── Generowanie error_details.txt (błąd krytyczny) ───────────────────────────

def _generate_critical_error_file(run_date: str, pipeline_run_id: int, conn) -> str | None:
    """
    Generuje error_details_{date}.txt z pełnym opisem błędu krytycznego.
    Jeśli nie było błędów krytycznych — zwraca None (plik nie jest tworzony).
    """
    columns, critical_errors = _safe_fetch(conn,
        """
        SELECT step_name, start_time, end_time, error_details
        FROM maintenance.pipeline_run_step
        WHERE pipeline_run_id = %s
          AND status = 'ERROR'
          AND step_name NOT LIKE 'scrape_%%'
        ORDER BY start_time;
        """,
        (pipeline_run_id,)
    )

    # Brak błędów krytycznych — nie twórz pliku
    if not critical_errors:
        print("[report] Brak błędów krytycznych — nie tworzę error_details.txt")
        return None

    file_path = os.path.join(REPORTS_DIR, f"error_details_{run_date}.txt")

    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f"=== BŁĘDY KRYTYCZNE PIPELINE: {run_date} ===\n\n")
        for row in critical_errors:
            if not row or len(row) < 4:
                continue
            step_name, start_time, end_time, error_details = row
            f.write(f"{'─' * 60}\n")
            f.write(f"Krok:        {step_name}\n")
            f.write(f"Start:       {start_time}\n")
            f.write(f"Koniec:      {end_time}\n\n")
            f.write("=== SZCZEGÓŁY BŁĘDU ===\n")
            f.write(error_details or "Brak szczegółów.")
            f.write("\n\n")

    print(f"[report] Zapisano błędy krytyczne: {file_path}")
    return file_path


# ── Generowanie error_details.txt (błąd dbt test) ───────────────────────────
def _generate_dbt_test_error_file(run_date: str, pipeline_run_id: int, conn) -> str | None:
    """
    Generuje error_dbt_test_silver_{date}.txt jeśli dbt_test_silver zakończył się błędem.
    """
    columns, rows = _safe_fetch(conn,
        """
        SELECT step_name, start_time, end_time, error_details
        FROM maintenance.pipeline_run_step
        WHERE pipeline_run_id = %s
          AND step_name = 'dbt_test_silver'
          AND status = 'ERROR';
        """,
        (pipeline_run_id,)
    )

    if not rows:
        print("[report] Brak błędów dbt test Silver — nie tworzę pliku txt.")
        return None

    file_path = os.path.join(REPORTS_DIR, f"error_dbt_test_silver_{run_date}.txt")

    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f"=== BŁĄD dbt test Silver: {run_date} ===\n\n")
        for row in rows:
            if not row or len(row) < 4:
                continue
            step_name, start_time, end_time, error_details = row
            f.write(f"{'─' * 60}\n")
            f.write(f"Krok:        {step_name}\n")
            f.write(f"Start:       {start_time}\n")
            f.write(f"Koniec:      {end_time}\n\n")
            f.write("=== SZCZEGÓŁY BŁĘDU ===\n")
            f.write(error_details or "Brak szczegółów.")
            f.write("\n\n")

    print(f"[report] Zapisano błędy dbt test: {file_path}")
    return file_path


# ── Wypełnienie notification_queue ───────────────────────────────────────────

def _populate_notification_queue(run_date: str, pipeline_run_id: int,
                                  pipeline_status: str, conn,
                                  failed_scrapers: list):
    """Wypełnia maintenance.notification_queue rekordami dla raportu."""
    records = []

    # MAIN_INFO — status pipeline
    records.append((
        pipeline_run_id, 'MAIN_INFO', 'PIPELINE',
        f"Status pipeline: {pipeline_status}",
        f"Pipeline praca IT wykonany dnia {run_date} — {pipeline_status}",
        False, None, None
    ))

    # MAIN_INFO — wgrane dane per portal
    records.append((
        pipeline_run_id, 'MAIN_INFO', 'PORTAL',
        'Wgrane dane per portal',
        f"Szczegóły wgranych danych dla daty {run_date}",
        False, None, None
    ))

    # MAIN_INFO — przestrzeń dyskowa
    records.append((
        pipeline_run_id, 'MAIN_INFO', 'STORAGE',
        'Przestrzeń dyskowa',
        "Zestawienie zajmowanej przestrzeni dyskowej",
        False, None, None
    ))

    # WARNING — failed scrapery (tylko jeśli były)
    for step_name in failed_scrapers:
        portal = step_name.replace("scrape_", "")
        records.append((
            pipeline_run_id, 'WARNING', 'PORTAL',
            f"Scraper {portal} nie powiódł się",
            f"Portal {portal} nie wgrał żadnych danych — szczegóły w załączniku txt",
            False, None, None
        ))

    # WARNING — widoki data quality (attach=true, per widok)
    for schema, view_name, tab_title in WARNING_VIEWS:
        records.append((
            pipeline_run_id, 'WARNING', 'DATA_QUALITY',
            tab_title,
            f"Widok {schema}.{view_name}",
            True, view_name, schema
        ))

    # ADD_INFO — szczegóły kroków pipeline
    records.append((
        pipeline_run_id, 'ADD_INFO', 'PIPELINE',
        'Szczegóły kroków pipeline',
        'Czasy i statusy poszczególnych kroków pipeline',
        True, 'pipeline_run_step', 'maintenance'
    ))

    # ADD_INFO — przestrzeń dyskowa szczegóły
    records.append((
        pipeline_run_id, 'ADD_INFO', 'STORAGE',
        'Przestrzeń dyskowa szczegóły',
        'Szczegółowe zestawienie przestrzeni dyskowej per schema',
        True, 'v_total_obj_sizes', 'maintenance'
    ))

    try:
        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO maintenance.notification_queue
                    (pipeline_run_id, type, category, title, message, attach, view_name, view_schema)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
                """,
                records
            )
            conn.commit()
        print(f"[report] Wstawiono {len(records)} rekordów do notification_queue")
    except Exception as e:
        conn.rollback()
        print(f"[report] BŁĄD zapisu do notification_queue: {e}")


# ── Główna funkcja task prepare_report ───────────────────────────────────────

def prepare_report(**context):
    """
    Task przygotowania raportu — uruchamiany zawsze (trigger_rule=ALL_DONE).
    1. Czyści folder data/reports/
    2. Robi TRUNCATE maintenance.notification_queue
    3. Określa status pipeline na podstawie kroków (nie kolumny status)
    4. Generuje pliki Excel i txt (tylko gdy są dane/błędy)
    5. Wypełnia notification_queue
    6. Zapisuje ścieżki plików do XCom dla t_send_report
    """
    run_date        = context['ti'].xcom_pull(task_ids='set_run_date')
    pipeline_run_id = context['ti'].xcom_pull(task_ids='set_run_date', key='pipeline_run_id')

    print(f"[report] Przygotowanie raportu dla daty {run_date}, pipeline_run_id={pipeline_run_id}")

    # Krok 1 — wyczyść folder reports
    _cleanup_reports_dir()

    conn = get_db_connection()
    try:
        # Krok 2 — TRUNCATE notification_queue
        try:
            with conn.cursor() as cur:
                cur.execute("TRUNCATE maintenance.notification_queue;")
                conn.commit()
            print("[report] TRUNCATE notification_queue — OK")
        except Exception as e:
            conn.rollback()
            print(f"[report] BŁĄD TRUNCATE notification_queue: {e}")

        # Krok 3 — określ status pipeline na podstawie kroków
        # Nie polegamy na kolumnie status w pipeline_run bo finalize może jeszcze nie działać

        critical_errors = _safe_fetch_scalar(conn,
            """
            SELECT COUNT(*) FROM maintenance.pipeline_run_step
            WHERE pipeline_run_id = %s
              AND status = 'ERROR'
              AND step_name NOT LIKE 'scrape_%%';
            """,
            (pipeline_run_id,), default=0
        )

        scraper_errors = _safe_fetch_scalar(conn,
            """
            SELECT COUNT(*) FROM maintenance.pipeline_run_step
            WHERE pipeline_run_id = %s
              AND step_name LIKE 'scrape_%%'
              AND status = 'ERROR';
            """,
            (pipeline_run_id,), default=0
        )

        total_scrapers = _safe_fetch_scalar(conn,
            """
            SELECT COUNT(*) FROM maintenance.pipeline_run_step
            WHERE pipeline_run_id = %s
              AND step_name LIKE 'scrape_%%';
            """,
            (pipeline_run_id,), default=0
        )

        # ERROR gdy: błąd krytyczny LUB wszystkie scrapery failed
        if critical_errors > 0 or (total_scrapers > 0 and scraper_errors == total_scrapers):
            pipeline_status = 'ERROR'
        elif scraper_errors > 0:
            pipeline_status = 'PARTIAL_SUCCESS'
        else:
            pipeline_status = 'SUCCESS'

        print(f"[report] Status pipeline: {pipeline_status} "
              f"(critical_errors={critical_errors}, "
              f"scraper_errors={scraper_errors}/{total_scrapers})")

        # Krok 4 — znajdź failed scrapery
        _, failed_rows = _safe_fetch(conn,
            """
            SELECT step_name FROM maintenance.pipeline_run_step
            WHERE pipeline_run_id = %s
              AND step_name LIKE 'scrape_%%'
              AND status = 'ERROR'
            ORDER BY step_name;
            """,
            (pipeline_run_id,)
        )
        failed_scrapers = [row[0] for row in failed_rows if row and len(row) > 0]

        if failed_scrapers:
            print(f"[report] Failed scrapery: {failed_scrapers}")
        else:
            print("[report] Brak failed scraperów.")

        # Krok 5 — generuj pliki (każda funkcja sama decyduje czy tworzyć plik)
        warnings_file       = _generate_warnings_excel(run_date, conn)
        details_file        = _generate_details_excel(run_date, pipeline_run_id, conn)
        scraper_error_files = _generate_scraper_error_files(run_date, pipeline_run_id, conn)
        dbt_test_error_file = _generate_dbt_test_error_file(run_date, pipeline_run_id, conn)
        critical_error_file = _generate_critical_error_file(run_date, pipeline_run_id, conn)

        # Krok 6 — wypełnij notification_queue
        _populate_notification_queue(
            run_date, pipeline_run_id, pipeline_status, conn, failed_scrapers
        )

        # Zapisz ścieżki plików do XCom dla t_send_report
        context['ti'].xcom_push(key='warnings_file',        value=warnings_file)
        context['ti'].xcom_push(key='details_file',         value=details_file)
        context['ti'].xcom_push(key='scraper_error_files',  value=scraper_error_files)
        context['ti'].xcom_push(key='dbt_test_error_file',  value=dbt_test_error_file)
        context['ti'].xcom_push(key='critical_error_file',  value=critical_error_file)
        context['ti'].xcom_push(key='pipeline_status',      value=pipeline_status)
        context['ti'].xcom_push(key='failed_scrapers',      value=failed_scrapers)

        print("[report] Przygotowanie raportu zakończone.")

    finally:
        conn.close()